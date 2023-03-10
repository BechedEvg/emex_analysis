import os
from operator import itemgetter
from time import sleep
import requests
import urllib3
import ssl
from openpyxl import load_workbook, Workbook
import pandas as pd


class CustomHttpAdapter(requests.adapters.HTTPAdapter):
    # "Transport adapter" that allows us to use custom ssl_context.

    def __init__(self, ssl_context=None, **kwargs):
        self.ssl_context = ssl_context
        super().__init__(**kwargs)

    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = urllib3.poolmanager.PoolManager(
            num_pools=connections, maxsize=maxsize,
            block=block, ssl_context=self.ssl_context)


def get_legacy_session():
    ctx = ssl.create_default_context(ssl.Purpose.SERVER_AUTH)
    ctx.options |= 0x4  # OP_LEGACY_SERVER_CONNECT
    session = requests.session()
    session.mount('https://', CustomHttpAdapter(ctx))
    return session


# Reading a list of elements from a source file.
class Exel_RW:

    def read_exel(file_name, count=-1):
        workbook = pd.read_excel(file_name)
        list_product = []
        for elements in workbook.values:
            list_product.append(list(elements)[:count])
        return list_product

    def write_exel(write_lists, file_name, sheet_name=0):
        if file_name not in os.listdir():
            workbook = Workbook()
            workbook.save(file_name)
            workbook.close()
        workbook = load_workbook(file_name)
        if sheet_name == 0:
            worksheet = workbook[workbook.sheetnames[0]]
        elif sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[sheet_name]
        for list_values in write_lists:
            worksheet.append(list_values)
        workbook.save(file_name)
        workbook.close()


class ProcessingList:

    def __init__(self, lists_data):
        self.lists_data = lists_data

    def data_processing(self):
        list_processing_data = []
        filter = True
        for values in self.lists_data:

            rating = (values[10]).replace(",", ".").replace("—", "0")
            quantity = (values[11])
            if type(quantity) == str:
                quantity = 5
            delivery = (values[12])

            if float(rating) < 4.5:
                filter = False
            if quantity < 3:
                filter = False
            if delivery > 5:
                filter = False
            if filter:
                list_processing_data.append(values)
        return list_processing_data

    def discount_calculation(self):
        list_processing_data = []
        dict_product_group = {
            "амортизаторы": 15,
            "водяные насосы": 12,
            "комплекты грм": 12,
            "модули и катушки зажигания": 13,
            "привод грм и агрегатов": 11,
            "свечи зажигания": 11,
            "ступицы и ступичные подшипники": 15,
            "термостаты": 12,
            "тормозная система": 12,
            "фильтры": 12,
            "шестерни грм": 12,
            "элементы подвески и рулевого управления": 13
        }
        for data in self.lists_data:
            price = data[8]
            if data[3].lower in dict_product_group:
                discount_percentage = dict_product_group[data[3].lower]
                price_discount = round(price - (price / 100 * discount_percentage))
            else:
                price_discount = round(price - (price / 100 * 13))
            list_processing_data.append(data[:9] + [price_discount] + data[9:])
        return list_processing_data


def get_html(url):
    html = get_legacy_session().get(url)
    return html


def get_lists_dict_analogs(dict_product):
    lists_dict_analogs_completed = []
    list_dict_analogs = dict_product["analogs"]
    for analog_dict in list_dict_analogs:
        offers = analog_dict['offers']
        for offer in offers:
            lists_dict_analogs_completed.append({
                "vendor_cod": analog_dict["detailNum"],
                "make": analog_dict['make'],
                "name": analog_dict['name'],
                "price": offer['displayPrice']['value'],
                "rating": offer['rating2']['rating'],
                "quantity": offer['quantity'],
                "delivery": offer['delivery']['value']
            })
    return lists_dict_analogs_completed


def get_lists_product(input_lists):
    write_list = []
    for list_product in input_lists:
        list_original_product = list_product[:5]
        vendor_cod = list_product[0]
        dict_product = get_emex_dict_products(vendor_cod)
        lists_dict_analogs = get_lists_dict_analogs(dict_product)
        for dict_analog in lists_dict_analogs:
            if dict_analog["quantity"] == 1000:
                dict_analog["quantity"] = "под заказ"
            write_list.append(list_original_product +
                              [dict_analog['vendor_cod'],
                               dict_analog['make'],
                               dict_analog['name'],
                               dict_analog["price"],
                               dict_analog["rating"],
                               dict_analog["quantity"],
                               dict_analog["delivery"],
                               f"https://emex.ru/products/{dict_analog['vendor_cod']}/{dict_analog['make']}/29241"])
    return write_list


def get_emex_dict_products(vendor_cod):
    url_part1 = "https://emex.ru/api/search/search?detailNum="
    url_part2 = "&locationId=29241&showAll=true"
    url = url_part1 + vendor_cod + url_part2
    dict_product = (get_html(url).json().get('searchResult'))
    return dict_product


def write_list_data(lists_product):
    column_names = [["Артикул OEM",
                     "Производитель OEM",
                     "Артикул DFR",
                     "Группа продукта",
                     "Наименование детали",
                     "Артикул аналога",
                     "Бренд аналога",
                     "Наименование аналога",
                     "Цена",
                     "Цена с учетом скидки",
                     "Рейтинг",
                     "Наличие, шт.",
                     "Срок доставки, дней",
                     "Ссылка"]]
    list_product = sorted(lists_product, key=itemgetter(0))
    list_product = sorted(list_product, key=itemgetter(5))
    list_product = sorted(list_product, key=itemgetter(9))
    return column_names + list_product


def write_list_analysis(dict_write):
    list_write_analysis = []
    for vendor_cod in dict_write:
        rez = dict_write[vendor_cod]["data"]
        brands = dict_write[vendor_cod]["brands"]
        list_write_analysis.append(rez + list(brands.values()))
    return list_write_analysis


def analysis(lists_data):
    list_analysis = [["Артикул OEM",
                      "Производитель OEM",
                      "Артикул DFR",
                      "Группа продукта",
                      "Наименование детали"]]
    dict_brands = get_dict_brand(lists_data)
    list_brands = list(dict_brands)
    list_analysis[0] += list_brands
    dict_write = {}
    for data in lists_data:
        if data[0] not in dict_write:
            dict_write[data[0]] = {}
            dict_write[data[0]]["data"] = data[:5]
            dict_write[data[0]]["brands"] = dict_brands.copy()
        if data[5] not in dict_write[data[0]]:
            dict_write[data[0]][data[5]] = []
        dict_write[data[0]][data[5]].append(data)

    for vendor_cod in dict_write:
        dict_write[vendor_cod]["brands"] = dict_brands.copy()
        for analog_cod in dict_write[vendor_cod]:
            if analog_cod != "brands" and analog_cod != "data":
                dict_write[vendor_cod][analog_cod] = sorted(dict_write[vendor_cod][analog_cod], key=itemgetter(9))[0]
                brand_price = dict_write[vendor_cod]["brands"][dict_write[vendor_cod][analog_cod][6]]
                if brand_price != "":
                    if brand_price > dict_write[vendor_cod][analog_cod][9]:
                        dict_write[vendor_cod]["brands"][dict_write[vendor_cod][analog_cod][6]] = \
                        dict_write[vendor_cod][analog_cod][9]
                else:
                    dict_write[vendor_cod]["brands"][dict_write[vendor_cod][analog_cod][6]] = \
                    dict_write[vendor_cod][analog_cod][9]
    list_analysis += write_list_analysis(dict_write)
    return list_analysis


def get_dict_brand(list_data):
    brands = []
    for list_brand in list_data:
        brands.append(list_brand[6])
    brands = sorted(set(brands))
    dict_brands = {}
    for brand in brands:
        dict_brands[brand] = ""
    return dict_brands


def main():
    input_list = Exel_RW.read_exel("input.xlsx")

    product_lists = get_lists_product(input_list)
    product_lists_discount = ProcessingList(product_lists).discount_calculation()
    product_lists = ProcessingList(product_lists_discount).data_processing()
    write_list_analysis = analysis(product_lists)

    write_lists_data = write_list_data(product_lists_discount)
    Exel_RW.write_exel(write_lists_data, "data.xlsx")
    Exel_RW.write_exel(write_list_analysis, "data.xlsx", "Sheet1")


if __name__ == '__main__':
    main()
